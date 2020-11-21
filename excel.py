import openpyxl
import nltk
import os
import time
import pickle

#----------------Génération des documents--------------
'''Cette fonction permet de générer de nouveau documents pour notre appretissage à partir des description des codes CIM '''
def generateCIM(excelFile):
    #Ouverture du fichier excel contenant les codes CIM
    cim_codes = openpyxl.load_workbook(excelFile)

    #La feuil contenant les codes et leur description 
    sheet = cim_codes.get_sheet_by_name('Feuil1')
    #Initialisations
    #Nombre de fichier traité (et sera aussi le nom de fichier généré n°fichier.txt )
    file_count = 0
    #dictionnaire qui va associer le code CIM à chaque fichier
    codes = nltk.defaultdict(str)

    #Création d'un répertoire qui va contenir les données
    if not os.path.exists("CIM_Dataset"):
        os.makedirs("CIM_Dataset")

    start_time = time.time()        
    #Lecture à partir du fichier
    for i in range (2,sheet.max_row+1):
        #print (sheet.cell(row = i, column = 2).value,sheet.cell(row = i, column = 3).value)
        #Géneration du fichier 
        fileName = "CIM_Dataset/C"+str(file_count)+'.txt'
        
        #Lecture du code et sa description
        code_description = str(sheet.cell(row = i, column = 3).value)
        cim_code = str(sheet.cell(row = i, column = 2).value)
        codes ['C'+str(file_count)] = cim_code
        # ouvrir le fichier de sortie en mode écriture
        file = open(fileName,'w', encoding = "utf-8")
        file.write(code_description)
        file.close()
        file_count += 1
    print("--- Temps d'exécution %s seconds ---" % (time.time() - start_time))
    pickle.dump( codes, open( "CIM_Dataset_Codes.p", "wb" ) )
    #print( codes )
    return codes

        
#---------------Code CIM des documents------------
'''Lire à partire des fichier excel les codes CIM associé à chaque diagnostic'''         
def getCIM(excelFile):
    #Ouverture du fichier excel contenant les code de chaque diagnostic
    cim_codes = openpyxl.load_workbook(excelFile)
    #Dictionnaire qui va associer à chaque diagnostique son code CIM
    codes = nltk.defaultdict(str)
    #Ouverture de la page Excel contentant les codes associés au diagnostics
    for sheet_name in ['CRH','CRO']:
        sheet = cim_codes.get_sheet_by_name(sheet_name)    
        for i in range (3,sheet.max_row+1+1):
            file_name = str(sheet.cell(row = i, column = 1).value)
            cim_code = sheet.cell(row = i, column = 2).value
            if cim_code != '/':
                codes [file_name] = cim_code
    pickle.dump( codes, open( "Diagnostics_Dataset_CIM_Codes.p", "wb" ) )
   # print (codes)
    return codes
    
#getCIM('Diagnostics.xlsx')


#--------------------------------------------------------------NGAP-------------------------------------

#----------------Génération des documents--------------
'''Cette fonction permet de générer de nouveau documents pour notre appretissage à partir des description des codes NGAP'''
def generateNGAP(excelFile):
    #Ouverture du fichier excel contenant les codes NGAP
    ngap_codes = openpyxl.load_workbook(excelFile)

    #La feuil contenant les codes et leur description 
    sheet = ngap_codes.get_sheet_by_name('Feuil1')
    #Initialisations
    #Nombre de fichier traité (et sera aussi le nom de fichier généré n°fichier.txt )
    file_count = 0
    #dictionnaire qui va associer le code NGAP à chaque fichier
    codes = nltk.defaultdict(str)

    #Création d'un répertoire qui va contenir les données
    if not os.path.exists("NGAP_Dataset"):
        os.makedirs("NGAP_Dataset")

    start_time = time.time()        
    #Lecture à partir du fichier
    for i in range (2,sheet.max_row+1):
       # print (sheet.cell(row = i, column = 2).value,sheet.cell(row = i, column = 3).value)
        #Géneration du fichier 
        fileName = "NGAP_Dataset/"+str(file_count)+'.txt'
        
        #Lecture du code et sa description
        code_description = str(sheet.cell(row = i, column = 2).value)
        cim_code = str(sheet.cell(row = i, column = 1).value)
        codes ['N'+str(file_count)] = cim_code
        #print (cim_code,code_description)
        # ouvrir le fichier de sortie en mode écriture
        file = open(fileName,'w', encoding = "utf-8")
        file.write(code_description)
        file.close()
        file_count += 1
    print("--- Temps d'exécution %s seconds ---" % (time.time() - start_time))
    pickle.dump( codes, open( "NGAP_Dataset_Codes.p", "wb" ) )
    #print( codes )
    return codes

        
#---------------Code NGAP des documents------------
'''Lire à partire des fichier excel les codes NGAP associé à chaque diagnostic'''         
def getNGAP(excelFile):
    #Ouverture du fichier excel contenant les code de chaque diagnostic
    cim_codes = openpyxl.load_workbook(excelFile)
    #Dictionnaire qui va associer à chaque diagnostique son code CIM
    codes = nltk.defaultdict(str)
    #Ouverture de la page Excel contentant les codes associés au diagnostics
    for sheet_name in ['CRH','CRO']:
        sheet = cim_codes.get_sheet_by_name(sheet_name)    
        for i in range (3,sheet.max_row+1+1):
            file_name = str(sheet.cell(row = i, column = 1).value)
            cim_code = sheet.cell(row = i, column = 4).value
            if cim_code != '/':
                codes [file_name] = cim_code
    pickle.dump( codes, open( "Diagnostics_Dataset_NGAP_Codes.p", "wb" ) )
   # print (codes)
    return codes
    
generateNGAP('NGAP 2006.xlsx')
generateCIM('CIM 10.xlsx')
getNGAP('Diagnostics.xlsx')
getCIM('Diagnostics.xlsx')







